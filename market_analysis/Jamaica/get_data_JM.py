"""
JaCars.net Car Listings Scraper  (Playwright version)
======================================================
Uses a real headless browser to bypass Cloudflare / bot protection.

Requirements:
    pip install playwright beautifulsoup4 openpyxl
    playwright install chromium

Usage:
    python get_data_JM.py
    Or: python -m market_analysis.Jamaica.get_data_JM
"""

import re
import time
import csv
from collections import Counter
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.sync_api import sync_playwright

BASE_URL = "https://www.jacars.net"

MANUFACTURERS = [
    ("Toyota",        "/cars/toyota/"),
    ("Honda",         "/cars/honda/"),
    ("Nissan",        "/cars/nissan/"),
    ("BMW",           "/cars/bmw/"),
    ("Subaru",        "/cars/subaru/"),
    ("Mercedes-Benz", "/cars/mercedes/"),
    ("Mazda",         "/cars/mazda/"),
    ("Suzuki",        "/cars/suzuki/"),
    ("Volkswagen",    "/cars/volkswagen/"),
    ("Hyundai",       "/cars/hyundai/"),
    ("KIA",           "/cars/kia/"),
    ("Audi",          "/cars/audi/"),
    ("Mitsubishi",    "/cars/mitsubishi/"),
    ("Ford",          "/cars/ford/"),
    ("Lexus",         "/cars/lexus/"),
    ("Land Rover",    "/cars/land-rover/"),
    ("Mini",          "/cars/mini/"),
    ("Daihatsu",      "/cars/daihatsu/"),
    ("Jaguar",        "/cars/jaguar/"),
    ("Porsche",       "/cars/porsche/"),
    ("Isuzu",         "/cars/isuzu/"),
    ("Infiniti",      "/cars/infiniti/"),
    ("Haval",         "/cars/haval/"),
    ("Chevrolet",     "/cars/chevrolet/"),
    ("Jeep",          "/cars/jeep/"),
    ("Acura",         "/cars/acura/"),
    ("Volvo",         "/cars/volvo/"),
    ("Peugeot",       "/cars/peugeot/"),
    ("BYD",           "/cars/byd/"),
    ("Chery",         "/cars/chery/"),
    ("Dodge",         "/cars/dodge/"),
    ("Chrysler",      "/cars/chrysler/"),
    ("GMC",           "/cars/gmc/"),
    ("Cadillac",      "/cars/cadillac/"),
    ("Tesla",         "/cars/tesla/"),
    ("Hino",          "/cars/hino/"),
    ("Fuso",          "/cars/fuso/"),
    ("UD",            "/cars/ud/"),
    ("Foton",         "/cars/foton/"),
    ("JAC",           "/cars/jac/"),
    ("Other",         "/cars/other/"),
]


# ── Parsing helpers ────────────────────────────────────────────────────────────

def parse_title(raw: str):
    """
    "Suzuki Swift RS 1,2L 2019"  ->  make="Suzuki", model_grade="Swift RS 1,2L", year="2019"
    Year is always the trailing 4-digit number.
    """
    raw = raw.strip()
    year_match = re.search(r'\b(19|20)\d{2}\b', raw)
    year = year_match.group(0) if year_match else ""
    remainder = raw[:year_match.start()].strip() if year_match else raw
    parts = remainder.split(None, 1)
    make = parts[0] if parts else ""
    model_grade = parts[1].strip() if len(parts) > 1 else ""
    return make, model_grade, year


def parse_price(item) -> str:
    """
    Extract current price, stripping the strikethrough discount span.
    Returns e.g. "JA$1,800,000"
    """
    price_el = item.select_one("a.advert__content-price span")
    if not price_el:
        return ""
    discount = price_el.find("span", class_=lambda c: c and "discount" in c)
    if discount:
        discount.decompose()
    return price_el.get_text(strip=True)


def parse_listings(html: str) -> list:
    soup = BeautifulSoup(html, "html.parser")
    listings = []
    for item in soup.select("div.advert.js-item-listing"):
        title_el = item.select_one("a.advert__content-title")
        if not title_el:
            continue
        raw_title = title_el.get_text(strip=True)
        href = title_el.get("href", "")
        full_url = BASE_URL + href if href.startswith("/") else href
        make, model_grade, year = parse_title(raw_title)
        price = parse_price(item)
        listings.append({
            "Year":        year,
            "Make":        make,
            "Model/Grade": model_grade,
            "Price (JA$)": price,
            "URL":         full_url,
        })
    return listings


def get_page_count(html: str) -> int:
    """Return the total number of pages from the pagination ul."""
    soup = BeautifulSoup(html, "html.parser")
    links = soup.select("ul.number-list a.page-number")
    if not links:
        return 1
    try:
        return max(int(a.get("data-page", 1)) for a in links)
    except ValueError:
        return 1


# ── Playwright scraper ─────────────────────────────────────────────────────────

def scrape_all():
    all_data = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            locale="en-US",
            viewport={"width": 1280, "height": 800},
        )
        page = context.new_page()

        # Warm up: visit homepage so cookies are set
        print("Visiting homepage to establish session...")
        page.goto(BASE_URL + "/", wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)

        for name, path in MANUFACTURERS:
            print(f"\nScraping {name}...")
            base_url = BASE_URL + path

            # Load page 1 to find total page count
            page.goto(base_url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(1.5)
            html = page.content()

            total_pages = get_page_count(html)
            print(f"  Total pages: {total_pages}")

            for pg in range(1, total_pages + 1):
                if pg == 1:
                    current_html = html
                else:
                    url = f"{base_url}?page={pg}"
                    print(f"  Page {pg}: {url}")
                    page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    time.sleep(1.5)
                    current_html = page.content()

                listings = parse_listings(current_html)
                # Fallback: use known manufacturer name if parse missed the make
                for lst in listings:
                    if not lst["Make"]:
                        lst["Make"] = name
                all_data.extend(listings)
                print(f"  Page {pg}: {len(listings)} listings (running total: {len(all_data)})")

        browser.close()

    return all_data


# ── Output ─────────────────────────────────────────────────────────────────────

def save_csv(all_data: list, filename: str = "jacars_listings.csv"):
    if not all_data:
        print("No data to save.")
        return
    fields = ["Year", "Make", "Model/Grade", "Price (JA$)", "URL"]
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(all_data)
    print(f"CSV saved -> {filename}")


def save_excel(all_data: list, filename: str = "jacars_listings.xlsx"):
    if not all_data:
        print("No data to save.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "All Listings"

    fields = ["Year", "Make", "Model/Grade", "Price (JA$)", "URL"]
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    row_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", fgColor="EEF2F7")

    for col, f in enumerate(fields, 1):
        c = ws.cell(row=1, column=col, value=f)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(all_data, 2):
        for col, f in enumerate(fields, 1):
            c = ws.cell(row=ri, column=col, value=row.get(f, ""))
            c.font = row_font
            if ri % 2 == 0:
                c.fill = alt_fill

    col_widths = {"Year": 8, "Make": 18, "Model/Grade": 30, "Price (JA$)": 18, "URL": 52}
    for col, f in enumerate(fields, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = col_widths[f]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:E1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary by Make")
    for col, val in enumerate(["Make", "Listings"], 1):
        c = ws2.cell(row=1, column=col, value=val)
        c.fill = hdr_fill
        c.font = hdr_font
    counts = Counter(r["Make"] for r in all_data)
    for ri, (make, count) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=ri, column=1, value=make).font = row_font
        ws2.cell(row=ri, column=2, value=count).font = row_font
        if ri % 2 == 0:
            ws2.cell(row=ri, column=1).fill = alt_fill
            ws2.cell(row=ri, column=2).fill = alt_fill
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12

    wb.save(filename)
    print(f"Excel saved -> {filename}")


def main():
    all_data = scrape_all()
    print(f"\nTotal listings scraped: {len(all_data)}")
    save_csv(all_data)
    save_excel(all_data)
    print("Done!")


if __name__ == "__main__":
    main()
